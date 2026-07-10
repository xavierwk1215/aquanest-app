# -*- coding: utf-8 -*-
"""
엑셀(fish_species_full.xlsx)을 수정한 뒤 다시 앱 데이터(JSON)로 변환하는 스크립트.

사용법:
    python3 reimport_from_excel.py <수정된 엑셀 경로>

3개 시트를 읽습니다:
  - "종 (Species)"           -> 종(species) 목록 (name/latin/genus/origin/maxSize/aliases/groupId/note)
  - "케어그룹 (CareGroups)"   -> 그룹(careGroups) 스펙 (temp/ph/tankMin/temperament/diet/... 등)
  - "합사예외 (PairOverrides)" -> SPECIES_PAIR_OVERRIDES 배열

주의: 종 시트의 [그룹] 접두어 붙은 컬럼들(temp_min 등)은 참고/열람용으로 같이 보여준 것이라
      실제로는 무시하고 "케어그룹" 시트의 값만 그룹 스펙으로 사용합니다.
      (여러 종이 같은 그룹을 공유하기 때문에, 종 시트에서 그룹 값을 고치면 어느 행을 기준으로
       삼아야 할지 애매해지는 문제를 피하기 위함입니다. 그룹 스펙을 고치고 싶으면
       반드시 "케어그룹" 시트에서 고쳐주세요.)
"""
import sys, json
from openpyxl import load_workbook

RESULT_MAP_KR_TO_EN = {"가능":"ok", "주의":"caution", "불가능":"bad"}

def yn(v):
    return str(v).strip().upper() == "Y"

def load_excel(path):
    wb = load_workbook(path, data_only=True)

    # ---- 케어그룹 시트 ----
    ws_g = wb["케어그룹 (CareGroups)"]
    headers_g = [c.value for c in ws_g[1]]
    groups = {}
    for row in ws_g.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        d = dict(zip(headers_g, row))
        gid = d["groupId"]
        groups[gid] = {
            "temp": [float(d["temp_min"]), float(d["temp_max"])],
            "ph": [float(d["ph_min"]), float(d["ph_max"])],
            "tankMin": int(d["tankMin_L"]),
            "temperament": d["temperament"],
            "waterSensitivity": d.get("waterSensitivity", "medium"),
            "finNipper": yn(d.get("finNipper","N")),
            "predatory": yn(d.get("predatory","N")),
            "sameSpeciesAggressionOnly": yn(d.get("sameSpeciesAggressionOnly","N")),
            "breedingAggressionOnly": yn(d.get("breedingAggressionOnly","N")),
            "territorial": yn(d.get("territorial","N")),
            "diet": d["diet"],
            "feeding": d["feeding"],
            "lifespan": d["lifespan"],
            "waterLevel": d["waterLevel"],
            "schooling": {"need": yn(d["schooling_need"]), "minGroup": int(d["schooling_minGroup"])},
            "difficulty": d["difficulty"],
            "color": "#38B6A3",  # 색상은 임시 기본값, 필요시 아래에서 temperament 기준으로 재계산
            "tips": d["tips"],
        }
        # temperament 기준으로 대략적인 색상 재계산 (peaceful=teal, semi=amber, aggressive=red)
        groups[gid]["color"] = {
            "peaceful": "#38B6A3", "semi-aggressive": "#F2A65A", "aggressive": "#E0714F"
        }.get(groups[gid]["temperament"], "#38B6A3")

    # ---- 종 시트 ----
    ws_s = wb["종 (Species)"]
    headers_s = [c.value for c in ws_s[1]]
    species_list = []
    for row in ws_s.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        d = dict(zip(headers_s, row))
        aliases_raw = d.get("aliases") or ""
        aliases = [a.strip() for a in str(aliases_raw).split(",") if a.strip()]
        species_list.append({
            "id": d["id"],
            "name": d["name"],
            "nameEn": d["nameEn"],
            "latin": d["latin"],
            "genus": d.get("genus") or "",
            "origin": d["origin"],
            "maxSize": float(d["maxSize_cm"]),
            "aliases": aliases,
            "groupId": d["groupId"],
            "note": (d.get("note") or "").strip() or None,
        })

    # ---- 합사예외 시트 ----
    overrides = []
    if "합사예외 (PairOverrides)" in wb.sheetnames:
        ws_o = wb["합사예외 (PairOverrides)"]
        headers_o = [c.value for c in ws_o[1]]
        for row in ws_o.iter_rows(min_row=3, values_only=True):
            if not row[0] or not row[2]:
                continue
            d = dict(zip(headers_o, row))
            result_kr = str(d.get("result","")).strip()
            result_en = RESULT_MAP_KR_TO_EN.get(result_kr)
            if not result_en:
                print(f"경고: 알 수 없는 result 값 건너뜀 -> {d}")
                continue
            overrides.append({
                "a": d["species_a_id"],
                "b": d["species_b_id"],
                "result": result_en,
                "reason": d.get("reason") or "",
                "source": d.get("source") or "",
            })

    return groups, species_list, overrides


def validate(groups, species_list):
    errors = []
    for gid, g in groups.items():
        if g["temp"][0] >= g["temp"][1]: errors.append(f"[{gid}] temp 범위 이상")
        if g["ph"][0] >= g["ph"][1]: errors.append(f"[{gid}] ph 범위 이상")
        if g["tankMin"] <= 0: errors.append(f"[{gid}] tankMin 이상")
        if g["temperament"] not in ("peaceful","semi-aggressive","aggressive"): errors.append(f"[{gid}] temperament 이상: {g['temperament']}")
        if g["difficulty"] not in ("초급","중급","고급"): errors.append(f"[{gid}] difficulty 이상: {g['difficulty']}")
    seen = set()
    for s in species_list:
        if s["id"] in seen: errors.append(f"중복 id: {s['id']}")
        seen.add(s["id"])
        if s["groupId"] not in groups: errors.append(f"[{s['id']}] 존재하지 않는 groupId: {s['groupId']}")
    return errors


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python3 reimport_from_excel.py <엑셀경로>")
        sys.exit(1)

    path = sys.argv[1]
    groups, species_list, overrides = load_excel(path)
    errors = validate(groups, species_list)

    print(f"그룹 {len(groups)}개, 종 {len(species_list)}개, 예외 {len(overrides)}건 불러옴")
    if errors:
        print(f"\n⚠ 검증 에러 {len(errors)}건:")
        for e in errors:
            print(" -", e)
    else:
        print("검증 통과: 에러 없음")

    out = {"groups": groups, "species": species_list, "overrides": overrides}
    with open("schema_reimported.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print("\nschema_reimported.json 저장 완료")
