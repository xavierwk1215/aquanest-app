# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

data = json.load(open("../data/species_data.json", encoding="utf-8"))
groups = data["groups"]
species_list = data["species"]

header_fill = PatternFill("solid", start_color="123A34", end_color="123A34")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
note_font = Font(italic=True, color="808080", name="Arial", size=8)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
cell_font = Font(name="Arial", size=10)

def style_header(ws, headers, note_row=True):
    for col, item in enumerate(headers, start=1):
        name, note = item if isinstance(item, tuple) else (item, "")
        c = ws.cell(row=1, column=col, value=name)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = 15
        if note_row:
            n = ws.cell(row=2, column=col, value=note)
            n.font = note_font
            n.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            n.border = border
    ws.row_dimensions[1].height = 28
    if note_row:
        ws.row_dimensions[2].height = 30

wb = Workbook()

# ---------------- Sheet 1: 종 (Species) ----------------
ws = wb.active
ws.title = "종 (Species)"
headers = [
    ("id","고유 ID (영문 하이픈) — 예외테이블에 이 값을 씁니다"),
    ("name","한글 이름"), ("nameEn","영문 이름"), ("latin","학명"), ("genus","속(genus)"),
    ("origin","원산지"), ("maxSize_cm","최대 크기(cm)"), ("aliases","유통명"),
    ("groupId","케어그룹 ID"), ("note","이 종만의 예외 사항"),
    ("temp_min","[그룹]수온 최소"), ("temp_max","[그룹]수온 최대"),
    ("ph_min","[그룹]pH 최소"), ("ph_max","[그룹]pH 최대"),
    ("tankMin_L","[그룹]최소 어항(L)"), ("temperament","[그룹]성격"),
    ("waterSensitivity","[그룹]수질예민도"),
    ("finNipper","[그룹]지느러미뭄"), ("predatory","[그룹]포식성"),
    ("sameSpeciesAggressionOnly","[그룹]동종간공격"), ("breedingAggressionOnly","[그룹]번식기공격"),
    ("territorial","[그룹]영역성"),
    ("diet","[그룹]먹이"), ("feeding","[그룹]급여빈도"), ("lifespan","[그룹]수명"),
    ("waterLevel","[그룹]서식수층"), ("schooling_need","[그룹]군집필요"), ("schooling_minGroup","[그룹]최소무리수"),
    ("difficulty","[그룹]난이도"), ("tips","[그룹]사육팁"),
]
style_header(ws, headers)
ws.freeze_panes = "A3"

bool_cols = {}  # column letter -> field name, for dropdown targeting
col_names = [h[0] for h in headers]
for r, s in enumerate(sorted(species_list, key=lambda x: x["groupId"]), start=3):
    g = groups[s["groupId"]]
    row = [
        s["id"], s["name"], s["nameEn"], s["latin"], s.get("genus",""),
        s["origin"], s["maxSize"], ", ".join(s.get("aliases",[])), s["groupId"], s.get("note") or "",
        g["temp"][0], g["temp"][1], g["ph"][0], g["ph"][1], g["tankMin"], g["temperament"],
        g.get("waterSensitivity","medium"),
        "Y" if g.get("finNipper") else "N", "Y" if g.get("predatory") else "N",
        "Y" if g.get("sameSpeciesAggressionOnly") else "N", "Y" if g.get("breedingAggressionOnly") else "N",
        "Y" if g.get("territorial") else "N",
        g["diet"], g["feeding"], g["lifespan"], g["waterLevel"],
        "Y" if g["schooling"]["need"] else "N", g["schooling"]["minGroup"],
        g["difficulty"], g["tips"],
    ]
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border; cell.font = cell_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

last_row = 2 + len(species_list)
def col_letter(name):
    return ws.cell(row=1, column=col_names.index(name)+1).column_letter

dv_temper = DataValidation(type="list", formula1='"peaceful,semi-aggressive,aggressive"', allow_blank=True)
dv_sens = DataValidation(type="list", formula1='"low,medium,high"', allow_blank=True)
dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
dv_diff = DataValidation(type="list", formula1='"초급,중급,고급"', allow_blank=True)
for dv in (dv_temper, dv_sens, dv_yn, dv_diff):
    ws.add_data_validation(dv)
dv_temper.add(f"{col_letter('temperament')}3:{col_letter('temperament')}{last_row}")
dv_sens.add(f"{col_letter('waterSensitivity')}3:{col_letter('waterSensitivity')}{last_row}")
for ync in ["finNipper","predatory","sameSpeciesAggressionOnly","breedingAggressionOnly","territorial","schooling_need"]:
    dv_yn.add(f"{col_letter(ync)}3:{col_letter(ync)}{last_row}")
dv_diff.add(f"{col_letter('difficulty')}3:{col_letter('difficulty')}{last_row}")

# ---------------- Sheet 2: 케어그룹 (CareGroups) ----------------
ws2 = wb.create_sheet("케어그룹 (CareGroups)")
headers2 = [
    "groupId","temp_min","temp_max","ph_min","ph_max","tankMin_L","temperament","waterSensitivity",
    "finNipper","predatory","sameSpeciesAggressionOnly","breedingAggressionOnly","territorial",
    "diet","feeding","lifespan","waterLevel","schooling_need","schooling_minGroup",
    "difficulty","species_count","tips",
]
style_header(ws2, headers2, note_row=False)
ws2.freeze_panes = "A2"

species_count = {}
for s in species_list:
    species_count[s["groupId"]] = species_count.get(s["groupId"], 0) + 1

for r, (gid, g) in enumerate(sorted(groups.items()), start=2):
    row = [
        gid, g["temp"][0], g["temp"][1], g["ph"][0], g["ph"][1], g["tankMin"], g["temperament"],
        g.get("waterSensitivity","medium"),
        "Y" if g.get("finNipper") else "N", "Y" if g.get("predatory") else "N",
        "Y" if g.get("sameSpeciesAggressionOnly") else "N", "Y" if g.get("breedingAggressionOnly") else "N",
        "Y" if g.get("territorial") else "N",
        g["diet"], g["feeding"], g["lifespan"], g["waterLevel"],
        "Y" if g["schooling"]["need"] else "N", g["schooling"]["minGroup"],
        g["difficulty"], species_count.get(gid,0), g["tips"],
    ]
    for c, val in enumerate(row, start=1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.border = border; cell.font = cell_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# ---------------- Sheet 3: 합사예외 (PairOverrides) ----------------
ws3 = wb.create_sheet("합사예외 (PairOverrides)")
headers3 = [
    ("species_a_id","어종 A의 id (시트1 id열 참고)"),
    ("species_a_name","참고용 — 자동 대조 안 됨, 메모용"),
    ("species_b_id","어종 B의 id (시트1 id열 참고)"),
    ("species_b_name","참고용 — 자동 대조 안 됨, 메모용"),
    ("result","가능 / 주의 / 불가능"),
    ("reason","이 예외가 왜 적용되는지 설명"),
    ("source","도감 / 커뮤니티 / 직접관찰 등"),
]
style_header(ws3, headers3)
ws3.freeze_panes = "A3"

id_to_name = {s["id"]: s["name"] for s in species_list}
existing_overrides = [
    ("betta", "corydoras", "주의",
     "베타는 하층에서 조용히 지내는 코리도라스는 대체로 무시하는 편이라 실제로 널리 쓰이는 조합입니다. 다만 베타 개체 성격 차이가 커서 완전히 안전하다고 보긴 어렵습니다.",
     "커뮤니티 관찰"),
    ("red-belly-piranha-sp", "silver-dollar", "주의",
     "피라냐와 실버달러는 남미에서 실제로 서식지를 공유하는 조합으로, 실버달러의 빠른 유영 속도 덕분에 충분히 큰 수조에서는 종종 함께 사육됩니다. 그래도 위험이 완전히 없는 건 아니라 대형 수조와 충분한 무리가 전제됩니다.",
     "도감/커뮤니티"),
]
for r, (a, b, result, reason, source) in enumerate(existing_overrides, start=3):
    row = [a, id_to_name.get(a,""), b, id_to_name.get(b,""), result, reason, source]
    for c, val in enumerate(row, start=1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.border = border; cell.font = cell_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# empty rows for user to add more
for r in range(3+len(existing_overrides), 3+len(existing_overrides)+30):
    for c in range(1, len(headers3)+1):
        ws3.cell(row=r, column=c).border = border

dv_result = DataValidation(type="list", formula1='"가능,주의,불가능"', allow_blank=True)
ws3.add_data_validation(dv_result)
dv_result.add(f"E3:E{3+len(existing_overrides)+30}")

for col_letter_ in ["A","C","B","D","F","G"]:
    ws3.column_dimensions[col_letter_].width = 20
ws3.column_dimensions["F"].width = 50

wb.save("../data/fish_species_full.xlsx")
print("saved:", len(species_list), "species,", len(groups), "groups,", len(existing_overrides), "overrides")
